"""Build publication supplementary tables from current  source data."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_DIR = DATA_DIR / "source_data"
METADATA_DIR = DATA_DIR / "metadata"
SUPP_DIR = ROOT / "outputs"
OUT_XLSX = SUPP_DIR / "tables" / "Supplementary_Tables.xlsx"


README_ROWS = [
    ("ST1_DatasetRegistry", "Defines targeted biological contexts and why each dataset is included.", "Is the design broader than a single PBMC3k example without becoming another benchmark?"),
    ("ST2_DatasetLock", "Records publication-facing public data sources, sizes, metadata fields and study-use notes.", "Can each real-data panel be traced to a public dataset source without exposing internal working files?"),
    ("ST3_AnalysisObjects", "Records processed analysis inputs using publication-facing descriptions rather than local analysis-object filenames.", "What processed inputs feed the figures and diagnostics?"),
    ("ST4_MethodBackbone", "Defines the eight representative methods and four mechanism families.", "Why these eight methods and how do they map to the broader DR taxonomy?"),
    ("ST5_ClaimEvidence", "Panel-level claim-to-evidence matrix for Fig. 2-Fig. 7, version .", "Can every main figure panel be traced to source data and a scientific question?"),
    ("ST6_RobustnessDesign", "Predeclared dimensionality and upstream-representation robustness design.", "Does the manuscript avoid relying only on 2D visualizations?"),
    ("ST7_AxisAvailability", "Method-specific availability of robustness axes.", "Which method-axis combinations are supported by the intended analysis design?"),
    ("ST8_MechSimDesign", "Compact known-truth mechanism simulation design.", "Which controlled mechanisms are stress-tested?"),
    ("ST9_MechSimCompletion", "Completion summary for the eight-method mechanism simulation suite.", "Were all eight representative methods completed in the compact simulation suite?"),
    ("ST10_MechSimSummary", "Scenario-level mechanism simulation support summary.", "How does each mechanism scenario support or falsify claim axes?"),
    ("ST11_SourceDataIndex", "Publication-facing source-data index for supplementary and main-figure evidence tables.", "Can reviewers trace each evidence block to a table or source-data file without internal computer paths?"),
    ("ST12_IntegratedEvidence", "Integrated evidence layers by method and family.", "Which evidence layer is the limiting factor for each method or family?"),
    ("ST13_EvidenceBottleneck", "Method-level weakest evidence layer and support summary.", "Which recommendation claims are constrained by the weakest evidence layer?"),
    ("ST14_RobustnessSynthesis", "Integrated robustness synthesis for dimension and perturbation responses.", "Do practical recommendations remain conditional under dimension and perturbation stress?"),
    ("ST15_PanelExpansion", "v3.4 record of hard-standard panel replacements and 10-panel evidence coverage.", "Which panels were replaced or retained, and what unique source-data question does each answer?"),
]


PANEL_EXPANSION_ROWS = [
    ("Fig2", "h", "Do cells differ in detected-gene depth within each biological context?", "fig2_detected_gene_quantiles.csv", "empirical descriptive"),
    ("Fig2", "i", "Do total-count depths differ across contexts in a way that may affect sparsity and likelihood assumptions?", "fig2_total_count_depth.csv", "empirical descriptive"),
    ("Fig2", "j", "Which contexts combine rare-state burden with shallow per-cell detection?", "fig2_rare_label_burden.csv; fig2_cell_complexity_metrics.csv", "empirical descriptive"),
    ("Fig3", "i", "Which methods change most across biological contexts?", "fig3_method_context_metric_range.csv", "empirical diagnostic"),
    ("Fig3", "j", "Do local and global preservation trade off differently across datasets and mechanism families?", "fig3_local_global_tradeoff_by_context.csv", "empirical diagnostic"),
    ("Fig4", "i", "Which diagnostic components sit above or below their predeclared thresholds by mechanism family?", "fig4_threshold_margin_by_family_metric.csv", "empirical diagnostic"),
    ("Fig4", "j", "Do family-level gate summaries hide within-family method spread?", "fig4_family_internal_gate_range.csv", "empirical diagnostic"),
    ("Fig6", "h", "Which datasets remain below diagnostic thresholds under output-dimension changes?", "fig6_dimension_failure_by_dataset_metric.csv", "empirical robustness"),
    ("Fig7", "a", "Which claim gates fail most often?", "fig7_gate_failure_by_gate.csv", "empirical diagnostic"),
    ("Fig7", "b", "Which methods fail which claim gates?", "fig7_gate_failure_by_method.csv", "empirical diagnostic"),
    ("Fig7", "c", "Which biological contexts fail which claim gates?", "fig7_gate_failure_by_dataset.csv", "empirical diagnostic"),
    ("Fig7", "d", "Which mechanism families fail which claim gates?", "fig7_gate_failure_by_family.csv", "empirical diagnostic"),
    ("Fig7", "e", "Which diagnostic components drive gate failures?", "fig7_component_failure_by_metric.csv", "empirical diagnostic"),
    ("Fig7", "f", "Which components fall above or below thresholds by mechanism family?", "fig7_component_margin_by_family_metric.csv", "empirical diagnostic"),
    ("Fig7", "g", "Which worst-case robustness metrics fail by method?", "fig7_worst_failure_by_method_metric.csv", "empirical robustness"),
    ("Fig7", "h", "Which methods fail under output-dimension changes?", "fig7_dimension_failure_by_method_metric.csv", "empirical robustness"),
    ("Fig7", "i", "Which methods fail under dropout or noise stress?", "fig7_perturbation_failure_by_method.csv", "empirical robustness"),
    ("Fig7", "j", "Which known-truth claim axes fail by method?", "fig7_simulation_failure_by_method_claim_axis.csv", "simulation"),
]


TABLE_CAPTIONS = {
    "README": (
        "Supplementary Tables workbook guide. This sheet lists all supplementary-table sheets, their purpose and the reviewer question addressed by each table."
    ),
    "ST1_DatasetRegistry": (
        "Supplementary Table 1 | Targeted biological contexts used for claim-specific validation. "
        "This table defines the empirical and simulation contexts used to support the evidence framework, including the biological regime represented by each dataset and the interpretive role assigned to it."
    ),
    "ST2_DatasetLock": (
        "Supplementary Table 2 | Empirical data objects and public-source provenance. "
        "This table records the public data sources, dimensions, metadata fields and study-use notes used to generate the real-data analyses. Internal working-file locations are intentionally omitted from the publication-facing supplement."
    ),
    "ST3_AnalysisObjects": (
        "Supplementary Table 3 | Processed analysis inputs used for embeddings and metrics. "
        "This table documents the processed input type, subset strategy, dimensions and annotation fields used by the embedding, diagnostic and robustness analyses. Local analysis-object filenames are omitted because they are implementation artifacts rather than publication-level provenance."
    ),
    "ST4_MethodBackbone": (
        "Supplementary Table 4 | Eight-method representative backbone and mechanism-family mapping. "
        "This table defines the representative method panel, links each method to a mechanism family and states why the panel spans the broader 26-method catalogue without functioning as a new full benchmark."
    ),
    "ST5_ClaimEvidence": (
        "Supplementary Table 5 | Panel-level claim-to-evidence matrix for Figs. 2-7. "
        "This table records, for each main-figure panel, the scientific question, evidence class, source-data file, statistical or diagnostic definition and reviewer-risk assessment."
    ),
    "ST6_RobustnessDesign": (
        "Supplementary Table 6 | Predeclared dimensionality and upstream-representation robustness design. "
        "This table records output-dimension, upstream-PCA, trajectory-dimension and simulation-latent-dimension axes used to test whether conclusions depend on a single two-dimensional representation."
    ),
    "ST7_AxisAvailability": (
        "Supplementary Table 7 | Method-axis availability for robustness analyses. "
        "This table distinguishes implemented, meaningful method-axis combinations from unavailable combinations, preventing unsupported robustness analyses from being imputed."
    ),
    "ST8_MechSimDesign": (
        "Supplementary Table 8 | Compact mechanism simulation design. "
        "This table defines the six known-truth simulation scenarios, their mechanism targets, parameters and claim axes used for controlled stress testing."
    ),
    "ST9_MechSimCompletion": (
        "Supplementary Table 9 | Eight-method completion in the mechanism simulation suite. "
        "This table verifies the completion status and metric coverage for the eight representative methods in the compact simulation suite."
    ),
    "ST10_MechSimSummary": (
        "Supplementary Table 10 | Scenario-level mechanism simulation summary. "
        "This table summarizes scenario-specific metric rows, thresholds and support outcomes behind the simulation-derived evidence."
    ),
    "ST11_SourceDataIndex": (
        "Supplementary Table 11 | Source-data index. "
        "This table records publication-facing source-data and metadata resources by evidence block. It is an index for reviewer navigation and does not include local computer paths or temporary working-file locations."
    ),
    "ST12_IntegratedEvidence": (
        "Supplementary Table 12 | Integrated evidence layers. "
        "This table records method-level and family-level evidence support across diagnostic gates, biological anchors and mechanism simulations."
    ),
    "ST13_EvidenceBottleneck": (
        "Supplementary Table 13 | Evidence bottlenecks by method. "
        "This table identifies the weakest evidence layer, minimum support and mean support for each method, supporting the bottleneck-based interpretation rule."
    ),
    "ST14_RobustnessSynthesis": (
        "Supplementary Table 14 | Integrated robustness synthesis. "
        "This table combines dimension-response and perturbation-response summaries used by the practical recommendation synthesis."
    ),
    "ST15_PanelExpansion": (
        "Supplementary Table 15 | Main-figure hard-standard replacement audit. "
        "This table records evidence-bearing panels added or retained during hard-standard cleanup, including scientific question, source data and evidence type."
    ),
}


def read_csv(path: Path, sep: str = ",") -> pd.DataFrame:
    return pd.read_csv(path, sep=sep)


def build_integrated_evidence() -> pd.DataFrame:
    method = read_csv(SOURCE_DIR / "integrated_method_evidence_layers.csv")
    method.insert(0, "evidence_scope", "method")

    family = read_csv(SOURCE_DIR / "integrated_family_evidence_layers.csv")
    family.insert(0, "evidence_scope", "family")
    family.insert(1, "method", "")

    columns = ["evidence_scope", "method", "family", "support_fraction", "n_rows", "evidence_layer"]
    return pd.concat([method[columns], family[columns]], ignore_index=True)


def build_integrated_robustness() -> pd.DataFrame:
    columns = [
        "evidence_block",
        "method",
        "family",
        "metric",
        "dimension_delta",
        "perturbation",
        "failure_fraction",
        "n_rows",
        "weakest_evidence_layer",
        "minimum_support",
        "mean_support",
    ]

    dim = read_csv(SOURCE_DIR / "integrated_dimension_delta_by_method.csv")
    dim.insert(0, "evidence_block", "dimension_delta_by_method")

    perturb_method = read_csv(SOURCE_DIR / "integrated_perturbation_failure_by_method.csv")
    perturb_method.insert(0, "evidence_block", "perturbation_failure_by_method")

    perturb_metric = read_csv(SOURCE_DIR / "integrated_perturbation_failure_by_metric.csv")
    perturb_metric.insert(0, "evidence_block", "perturbation_failure_by_metric")
    perturb_metric.insert(1, "method", "")
    perturb_metric.insert(2, "family", "")

    bottleneck = read_csv(SOURCE_DIR / "integrated_method_evidence_bottleneck.csv")
    bottleneck.insert(0, "evidence_block", "mean_vs_minimum_support")
    bottleneck["metric"] = ""
    bottleneck["dimension_delta"] = ""
    bottleneck["perturbation"] = ""
    bottleneck["failure_fraction"] = ""
    bottleneck["n_rows"] = ""

    blocks = [dim, perturb_method, perturb_metric, bottleneck]
    normalized = []
    for block in blocks:
        for col in columns:
            if col not in block.columns:
                block[col] = ""
        normalized.append(block[columns])
    return pd.concat(normalized, ignore_index=True)


PUBLIC_DATA_SOURCE = {
    "pbmc3k": "Scanpy PBMC3k processed dataset",
    "paul15": "Scanpy Paul15 haematopoiesis dataset",
    "heart_cell_atlas_subsampled": "scvi-tools heart cell atlas subsampled dataset",
    "known_truth_count_simulation": "controlled known-truth simulation generated for this study",
}


def build_public_dataset_lock() -> pd.DataFrame:
    df = read_csv(SOURCE_DIR / "dataset_summary.csv")
    df = df.rename(columns={"loader": "public_data_source", "status": "study_use_status"})
    columns = [
        "dataset_id",
        "role",
        "study_use_status",
        "public_data_source",
        "n_obs",
        "n_vars",
        "label_candidates",
        "batch_candidates",
        "selected_label_field",
        "selected_batch_field",
        "min_label_count",
        "n_label_levels",
        "n_batch_levels",
        "notes",
    ]
    return df[[col for col in columns if col in df.columns]].copy()


def build_public_analysis_objects() -> pd.DataFrame:
    df = read_csv(SOURCE_DIR / "analysis_object_manifest.csv")
    df["public_data_source"] = df["dataset_id"].map(PUBLIC_DATA_SOURCE).fillna(df["dataset_id"])
    df["processed_input_description"] = (
        "processed expression matrix with highly variable genes, normalized representation and PCA reference used for embeddings and diagnostics"
    )
    columns = [
        "dataset_id",
        "public_data_source",
        "processed_input_description",
        "subset_strategy",
        "n_obs",
        "n_vars",
        "n_pcs",
        "label_field",
        "batch_field",
        "counts_integer_fraction",
    ]
    return df[[col for col in columns if col in df.columns]].copy()


def build_public_claim_evidence() -> pd.DataFrame:
    df = read_csv(METADATA_DIR / "main_figure_claim_to_evidence_matrix.tsv", sep="\t")
    columns = [
        "figure",
        "panel",
        "figure_main_claim",
        "panel_scientific_question",
        "panel_claim_type",
        "evidence_block",
        "unique_evidence_responsibility",
        "evidence_grade",
        "data_source_class",
        "evidence_data_type",
        "counts_toward_evidence_panel_target",
        "metric",
        "unit",
        "data_source",
        "sample_type",
        "n_definition",
        "independent_unit",
        "replicate_definition",
        "statistical_test_or_model",
        "multiple_testing_correction",
        "effect_size_definition",
        "confidence_interval_or_error_bar",
        "missing_value_handling",
        "exclusion_criteria",
        "batch_or_confounder_control",
        "reviewer_risk",
        "stress_test_status",
        "figure_decision",
        "panel_decision",
    ]
    out = df[[col for col in columns if col in df.columns]].copy()
    if "data_source" in out.columns:
        out["data_source"] = out["data_source"].map(publication_resource_label)
    out["publication_trace"] = out["figure"].astype(str).str.replace("Fig", "Source Data for Fig. ", regex=False)
    out["supplementary_trace"] = "Panel-level evidence definitions are summarized in Supplementary Table 5; public dataset sources are listed in Supplementary Tables 1-3."
    return out


def publication_resource_label(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    replacements = {
        "dataset_summary.csv": "empirical dataset summary in Supplementary Table 2",
        "dataset_composition.csv": "dataset composition summary derived from Supplementary Tables 1-3",
    }
    parts = []
    for part in text.split(";"):
        item = part.strip().replace("\\", "/")
        basename = item.split("/")[-1]
        if basename in replacements:
            parts.append(replacements[basename])
        elif basename.endswith(".csv"):
            readable = basename.removesuffix(".csv").replace("_", " ")
            if readable.startswith("fig"):
                parts.append(f"source-data table: {readable}")
            else:
                parts.append(readable)
        elif basename:
            parts.append(basename)
    return "; ".join(parts)


def build_source_data_index() -> pd.DataFrame:
    rows = [
        ("Fig. 2", "dataset context and pre-analysis constraints", "Supplementary Tables 1-3; Source Data for Fig. 2", "public single-cell datasets listed in Supplementary Tables 1-3"),
        ("Fig. 3", "matched empirical embeddings and diagnostic support", "Supplementary Tables 4-5; Source Data for Fig. 3", "processed public single-cell datasets listed in Supplementary Table 3"),
        ("Fig. 4", "claim-specific diagnostic gates", "Supplementary Tables 5-7; Source Data for Fig. 4", "diagnostic metrics derived from the representative eight-method panel"),
        ("Fig. 5", "independent biological-anchor audit", "Supplementary Tables 1-5; Source Data for Fig. 5", "marker, rare-state, continuum and donor-aware summaries derived from public datasets"),
        ("Fig. 6", "robustness to dimension, upstream representation and perturbation", "Supplementary Tables 6-7 and 14; Source Data for Fig. 6", "robustness metrics derived from predeclared perturbation and dimension-response analyses"),
        ("Fig. 7", "claim-specific failure synthesis", "Supplementary Tables 5-7 and 15; Source Data for Fig. 7", "diagnostic-gate, robustness and simulation failure summaries"),
        ("Supplementary Fig. S1", "dataset structure and analysis context", "Supplementary Tables 1-3", "publication-facing public dataset and processed-input summaries"),
        ("Supplementary Fig. S2", "metric-level mechanism simulation audit", "Supplementary Tables 8-10", "controlled known-truth simulations generated for this study"),
        ("Supplementary Fig. S3", "full embedding atlas", "Supplementary Tables 3-5", "embedding coordinates and labels generated from the representative eight-method panel"),
        ("Supplementary Fig. S4", "diagnostic-gate metric atlas", "Supplementary Tables 5-7", "diagnostic-gate source tables summarized in the workbook"),
        ("Supplementary Fig. S5", "independent biological anchors", "Supplementary Tables 1-5", "biological-anchor summaries derived from public single-cell datasets"),
        ("Supplementary Fig. S6", "robustness response atlas", "Supplementary Tables 6-7 and 14", "dimension-response, upstream-representation and perturbation summaries"),
        ("Supplementary Note 1", "unified mathematical formulations", "Supplementary Note 1", "method derivations and notation, provided as the accompanying supplementary note"),
        ("Supplementary Tables", "publication-facing tabular provenance", "this Excel workbook", "captions and legends are embedded in the first row of each sheet"),
    ]
    return pd.DataFrame(
        rows,
        columns=[
            "manuscript_item",
            "evidence_block",
            "publication_source_location",
            "traceability_note",
        ],
    )


def build_tables() -> dict[str, pd.DataFrame]:
    return {
        "README": pd.DataFrame(README_ROWS, columns=["sheet", "purpose", "reviewer_question"]),
        "ST1_DatasetRegistry": read_csv(METADATA_DIR / "dataset_registry.csv"),
        "ST2_DatasetLock": build_public_dataset_lock(),
        "ST3_AnalysisObjects": build_public_analysis_objects(),
        "ST4_MethodBackbone": read_csv(METADATA_DIR / "method_backbone.csv"),
        "ST5_ClaimEvidence": build_public_claim_evidence(),
        "ST6_RobustnessDesign": read_csv(METADATA_DIR / "dimension_response_design.csv"),
        "ST7_AxisAvailability": read_csv(SOURCE_DIR / "fig6_method_coverage.csv").rename(columns={"coverage_note": "availability_note"}),
        "ST8_MechSimDesign": read_csv(SOURCE_DIR / "fig6_mechanism_simulation_design.csv"),
        "ST9_MechSimCompletion": read_csv(SOURCE_DIR / "fig6_mechanism_simulation_coverage.csv").rename(columns={"coverage_note": "completion_note"}),
        "ST10_MechSimSummary": read_csv(SOURCE_DIR / "fig6_mechanism_simulation_scenario_summary.csv"),
        "ST11_SourceDataIndex": build_source_data_index(),
        "ST12_IntegratedEvidence": build_integrated_evidence(),
        "ST13_EvidenceBottleneck": read_csv(SOURCE_DIR / "integrated_method_evidence_bottleneck.csv"),
        "ST14_RobustnessSynthesis": build_integrated_robustness(),
        "ST15_PanelExpansion": pd.DataFrame(
            PANEL_EXPANSION_ROWS,
            columns=["figure", "panel", "new_scientific_question", "source_data_resource", "evidence_type"],
        ),
    }


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    caption_fill = PatternFill("solid", fgColor="F2F2F2")
    caption_font = Font(bold=True, color="000000", name="Times New Roman", size=10)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000", name="Arial", size=10)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    caption_alignment = Alignment(horizontal="justify", vertical="top", wrap_text=True)
    header_alignment = Alignment(vertical="top", wrap_text=True)

    for ws in wb.worksheets:
        caption = TABLE_CAPTIONS.get(ws.title, f"{ws.title}. This sheet provides supplementary tabular evidence for the manuscript.")
        ws.insert_rows(1)
        last_col = max(1, ws.max_column)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        caption_cell = ws.cell(row=1, column=1)
        caption_cell.value = caption
        caption_cell.fill = caption_fill
        caption_cell.font = caption_font
        caption_cell.alignment = caption_alignment
        ws.row_dimensions[1].height = 48
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.alignment = body_alignment
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(2, ws.max_row + 1)]
            max_len = max(len(str(value)) if value is not None else 0 for value in values)
            width = max(10, min(48, max_len + 2))
            ws.column_dimensions[column_letter].width = width
    wb.save(path)


def main() -> None:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    tables = build_tables()
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    style_workbook(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    for sheet_name, df in tables.items():
        print(f"{sheet_name}: {len(df)} data rows")


if __name__ == "__main__":
    main()


